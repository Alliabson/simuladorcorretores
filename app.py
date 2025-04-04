import streamlit as st
from datetime import datetime, timedelta
from PIL import Image
import locale
from math import ceil
from io import BytesIO
import os
import subprocess
import sys
import re  # Adicionando a importação faltante

# Configuração robusta do locale
def configure_locale():
    try:
        # Tenta configurar o locale específico do Brasil
        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
    except locale.Error:
        try:
            locale.setlocale(locale.LC_ALL, 'pt_BR')
        except locale.Error:
            try:
                locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
            except locale.Error:
                try:
                    locale.setlocale(locale.LC_ALL, '')
                except locale.Error:
                    # Fallback para locale neutro
                    locale.setlocale(locale.LC_ALL, 'C.UTF-8')
                    st.warning("Configuração de locale específica não disponível. Usando padrão internacional.")

# Configura o locale no início da execução
configure_locale()

# Instalação garantida de dependências
def install_and_import(package, import_name=None):
    import_name = import_name or package
    try:
        return __import__(import_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        return __import__(import_name)

pd = install_and_import('pandas')
np = install_and_import('numpy')
FPDF = install_and_import('fpdf2', 'fpdf').FPDF
npf = install_and_import('numpy-financial', 'numpy_financial')

@st.cache_data
def load_logo():
    try:
        # Tenta carregar a imagem localmente
        logo = Image.open("Celeste_Logo_Nova_2.png")
        return logo
    except Exception as e:
        st.warning(f"Não foi possível carregar a logo: {str(e)}")
        return None

#Executar em wide
st.set_page_config(layout="wide")

# Configuração do tema Streamlit
def set_theme():
    st.markdown("""
    <style>
        /* Fundo principal */
        .stApp {
            background-color: #1E1E1E;
        }
        
        /* Sidebar */
        [data-testid="stSidebar"] {
            background-color: #252526;
        }
        
        /* Títulos */
        h1, h2, h3, h4, h5, h6, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3 {
            color: #FFFFFF;
        }
        
        /* Texto geral */
        .stMarkdown p, .stMarkdown li, .stText, .stNumberInput label, .stSelectbox label {
            color: #E0E0E0;
        }
        
        /* Inputs */
        .stTextInput input, .stNumberInput input, .stSelectbox select {
            background-color: #333333;
            color: #FFFFFF;
            border-color: #555555;
        }
        
        /* Botões */
        .stButton button {
            background-color: #0056b3;
            color: white;
            border: none;
            border-radius: 4px;
        }
        
        .stButton button:hover {
            background-color: #003d82;
        }
        
        /* Cards/metricas */
        .stMetric {
            background-color: #252526;
            border-radius: 8px;
            padding: 15px;
            border-left: 4px solid #0056b3;
        }
        
        .stMetric label {
            color: #A0A0A0 !important;
        }
        
        .stMetric div {
            color: #FFFFFF !important;
            font-size: 24px !important;
        }
        
        /* Dataframe */
        .dataframe {
            background-color: #252526 !important;
            color: #E0E0E0 !important;
        }
        
        .dataframe th {
            background-color: #0056b3 !important;
            color: white !important;
        }
        
        .dataframe tr:nth-child(even) {
            background-color: #333333 !important;
        }
        
        .dataframe tr:hover {
            background-color: #444444 !important;
        }
    </style>
    """, unsafe_allow_html=True)

# Função de formatação de moeda robusta
def formatar_moeda(valor, simbolo=True):
    """Formata valores monetários com fallback manual"""
    try:
        if valor is None:
            return "R$ 0,00" if simbolo else "0,00"
        
        # Converte para float se for string
        if isinstance(valor, str):
            valor = re.sub(r'[^\d,]', '', valor).replace(',', '.')
            valor = float(valor)
        
        # Formatação manual independente do locale
        valor_abs = abs(valor)
        parte_inteira = int(valor_abs)
        parte_decimal = int(round((valor_abs - parte_inteira) * 100))
        
        # Formata parte inteira com separadores de milhar
        parte_inteira_str = f"{parte_inteira:,}".replace(",", ".")
        
        # Monta o resultado
        valor_formatado = f"{parte_inteira_str},{parte_decimal:02d}"
        
        # Adiciona sinal negativo se necessário
        if valor < 0:
            valor_formatado = f"-{valor_formatado}"
        
        # Adiciona símbolo se solicitado
        if simbolo:
            return f"R$ {valor_formatado}"
        return valor_formatado
    except Exception:
        return "R$ 0,00" if simbolo else "0,00"

def calcular_taxas(taxa_mensal):
    try:
        taxa_mensal_decimal = float(taxa_mensal) / 100
        taxa_anual = ((1 + taxa_mensal_decimal) ** 12) - 1
        taxa_semestral = ((1 + taxa_mensal_decimal) ** 6) - 1
        taxa_trimestral = ((1 + taxa_mensal_decimal) ** 3) - 1
        taxa_diaria = ((1 + taxa_mensal_decimal) ** (1/30)) - 1
        
        return {
            'anual': taxa_anual,
            'semestral': taxa_semestral,
            'trimestral': taxa_trimestral,
            'mensal': taxa_mensal_decimal,
            'diaria': taxa_diaria
        }
    except:
        return {
            'anual': 0,
            'semestral': 0,
            'trimestral': 0,
            'mensal': 0,
            'diaria': 0
        }

def calcular_valor_presente(valor_futuro, taxa, dias):
    try:
        if dias <= 0 or taxa <= 0:
            return valor_futuro
        return valor_futuro / ((1 + taxa) ** (dias / 30))
    except:
        return valor_futuro

def ajustar_data_vencimento(data, periodo, num_periodo=1, dia_vencimento=None):
    try:
        if not isinstance(data, datetime):
            data = datetime.combine(data, datetime.min.time())
            
        if periodo == "mensal":
            total_meses = data.month + num_periodo
            ano = data.year + (total_meses - 1) // 12
            mes = (total_meses - 1) % 12 + 1
            
            ultimo_dia_mes = (datetime(ano, mes + 1, 1) - timedelta(days=1)).day if mes < 12 else 31
            dia = min(dia_vencimento, ultimo_dia_mes) if dia_vencimento else ultimo_dia_mes
            
            return datetime(ano, mes, dia)
        
        elif periodo == "semestral":
            nova_data = data + timedelta(days=180 * num_periodo)
            ultimo_dia_mes = (datetime(nova_data.year, nova_data.month % 12 + 1, 1) - timedelta(days=1)).day
            dia = min(dia_vencimento, ultimo_dia_mes) if dia_vencimento else ultimo_dia_mes
            
            return datetime(nova_data.year, nova_data.month, dia)
        
        elif periodo == "anual":
            nova_data = data.replace(year=data.year + num_periodo)
            ultimo_dia_mes = (datetime(nova_data.year, nova_data.month % 12 + 1, 1) - timedelta(days=1)).day
            dia = min(dia_vencimento, ultimo_dia_mes) if dia_vencimento else ultimo_dia_mes
            
            return datetime(nova_data.year, nova_data.month, dia)
            
    except ValueError:
        return datetime(data.year, data.month, 28)

def parse_date(date_str):
    try:
        return datetime.strptime(date_str, '%d/%m/%Y')
    except ValueError:
        return datetime.now()

def calcular_comissoes(valor_total, comissao_coordenacao, comissao_imobiliaria):
    try:
        coord = (float(comissao_coordenacao) / 100) * float(valor_total)
        imob = (float(comissao_imobiliaria) / 100) * float(valor_total)
        return {
            'coord': coord,
            'imob': imob,
            'total': coord + imob
        }
    except:
        return {
            'coord': 0,
            'imob': 0,
            'total': 0
        }

def determinar_modo_calculo(modalidade):
    if modalidade == "mensal":
        return 1
    elif modalidade == "mensal + balão":
        return 2
    elif modalidade == "só balão anual":
        return 3
    elif modalidade == "só balão semestral":
        return 4
    else:
        return 1

def calcular_parcela(valor, taxa, periodos):
    try:
        if periodos <= 0 or taxa <= 0 or valor <= 0:
            return 0
            
        parcela = npf.pmt(taxa, periodos, -valor)
        
        if parcela is None or isinstance(parcela, str) or np.isnan(parcela) or np.isinf(parcela):
            return 0
            
        return abs(float(parcela))
    except Exception as e:
        st.error(f"Erro no cálculo da parcela: {str(e)}")
        return 0

def calcular_valor_presente_total(valor, taxa, periodos):
    try:
        if periodos <= 0 or taxa <= 0:
            return 0
            
        valor_presente = npf.pv(taxa, periodos, -valor)
        
        if np.isnan(valor_presente) or np.isinf(valor_presente):
            return 0
            
        return abs(float(valor_presente))
    except:
        return 0

def atualizar_baloes(modalidade, qtd_parcelas, tipo_balao=None):
    try:
        qtd_parcelas = int(qtd_parcelas)
        if modalidade == "mensal + balão":
            if tipo_balao == "anual":
                return max(qtd_parcelas // 12, 0)
            elif tipo_balao == "semestral":
                return max(qtd_parcelas // 6, 0)
        elif modalidade == "só balão anual":
            return max(ceil(qtd_parcelas / 12), 0)
        elif modalidade == "só balão semestral":
            return max(ceil(qtd_parcelas / 6), 0)
        return 0
    except:
        return 0

def gerar_cronograma(valor_financiado, valor_parcela, valor_balao,
                    qtd_parcelas, qtd_baloes, modalidade, tipo_balao,
                    data_entrada, taxas):
    cronograma = []
    try:
        if not isinstance(data_entrada, datetime):
            data_entrada = datetime.combine(data_entrada, datetime.min.time())
            
        saldo_devedor = float(valor_financiado)
        total_valor_presente = 0
        dia_vencimento = data_entrada.day
        
        parcelas = []
        baloes = []
        
        if modalidade == "mensal":
            for i in range(1, qtd_parcelas + 1):
                data_vencimento_parcela = ajustar_data_vencimento(data_entrada, "mensal", i, dia_vencimento)
                dias_corridos = (data_vencimento_parcela - data_entrada).days
                
                juros = saldo_devedor * taxas['mensal']
                amortizacao = valor_parcela - juros
                if amortizacao > saldo_devedor:
                    amortizacao = saldo_devedor
                    valor_parcela = amortizacao + juros
                
                saldo_devedor -= amortizacao
                
                valor_presente = calcular_valor_presente(valor_parcela, taxas['mensal'], dias_corridos)
                total_valor_presente += valor_presente
                
                parcelas.append({
                    "Item": f"Parcela {i}",
                    "Tipo": "Parcela",
                    "Data_Vencimento": data_vencimento_parcela.strftime('%d/%m/%Y'),
                    "Data_Pagamento": data_entrada.strftime('%d/%m/%Y'),
                    "Dias": dias_corridos,
                    "Valor": valor_parcela,
                    "Valor_Presente": valor_presente,
                    "Desconto_Aplicado": valor_parcela - valor_presente
                })
        
        elif modalidade in ["só balão anual", "só balão semestral"]:
            periodo = "anual" if modalidade == "só balão anual" else "semestral"
            taxa_periodo = taxas['anual'] if periodo == "anual" else taxas['semestral']
            
            for i in range(1, qtd_baloes + 1):
                data_vencimento_balao = ajustar_data_vencimento(data_entrada, periodo, i, dia_vencimento)
                dias_corridos = (data_vencimento_balao - data_entrada).days
                
                juros = saldo_devedor * taxa_periodo
                amortizacao = valor_balao - juros
                if amortizacao > saldo_devedor:
                    amortizacao = saldo_devedor
                    valor_balao = amortizacao + juros
                
                saldo_devedor -= amortizacao
                
                valor_presente = calcular_valor_presente(valor_balao, taxas['mensal'], dias_corridos)
                total_valor_presente += valor_presente
                
                baloes.append({
                    "Item": f"Balão {i}",
                    "Tipo": "Balão",
                    "Data_Vencimento": data_vencimento_balao.strftime('%d/%m/%Y'),
                    "Data_Pagamento": data_entrada.strftime('%d/%m/%Y'),
                    "Dias": dias_corridos,
                    "Valor": valor_balao,
                    "Valor_Presente": valor_presente,
                    "Desconto_Aplicado": valor_balao - valor_presente
                })
        
        elif modalidade == "mensal + balão":
            intervalo_balao = 12 if tipo_balao == "anual" else 6
            taxa_periodo = taxas['anual'] if tipo_balao == "anual" else taxas['semestral']
            
            contador_baloes = 1
            
            for i in range(1, qtd_parcelas + 1):
                data_vencimento = ajustar_data_vencimento(data_entrada, "mensal", i, dia_vencimento)
                dias_corridos = (data_vencimento - data_entrada).days
                
                juros_parcela = saldo_devedor * taxas['mensal']
                amortizacao_parcela = valor_parcela - juros_parcela
                if amortizacao_parcela > saldo_devedor:
                    amortizacao_parcela = saldo_devedor
                    valor_parcela = amortizacao_parcela + juros_parcela
                
                saldo_devedor -= amortizacao_parcela
                
                valor_presente_parcela = calcular_valor_presente(valor_parcela, taxas['mensal'], dias_corridos)
                total_valor_presente += valor_presente_parcela
                
                parcelas.append({
                    "Item": f"Parcela {i}",
                    "Tipo": "Parcela",
                    "Data_Vencimento": data_vencimento.strftime('%d/%m/%Y'),
                    "Data_Pagamento": data_entrada.strftime('%d/%m/%Y'),
                    "Dias": dias_corridos,
                    "Valor": valor_parcela,
                    "Valor_Presente": valor_presente_parcela,
                    "Desconto_Aplicado": valor_parcela - valor_presente_parcela
                })
                
                if i % intervalo_balao == 0 and contador_baloes <= qtd_baloes:
                    juros_balao = saldo_devedor * taxa_periodo
                    amortizacao_balao = valor_balao - juros_balao
                    if amortizacao_balao > saldo_devedor:
                        amortizacao_balao = saldo_devedor
                        valor_balao = amortizacao_balao + juros_balao
                    
                    saldo_devedor -= amortizacao_balao
                    
                    valor_presente_balao = calcular_valor_presente(valor_balao, taxas['mensal'], dias_corridos)
                    total_valor_presente += valor_presente_balao
                    
                    baloes.append({
                        "Item": f"Balão {contador_baloes}",
                        "Tipo": "Balão",
                        "Data_Vencimento": data_vencimento.strftime('%d/%m/%Y'),
                        "Data_Pagamento": data_entrada.strftime('%d/%m/%Y'),
                        "Dias": dias_corridos,
                        "Valor": valor_balao,
                        "Valor_Presente": valor_presente_balao,
                        "Desconto_Aplicado": valor_balao - valor_presente_balao
                    })
                    contador_baloes += 1
        
        cronograma = parcelas + baloes
        
        # Ajuste para garantir que o valor presente total seja igual ao valor financiado
        fator_ajuste = valor_financiado / total_valor_presente if total_valor_presente > 0 else 1
        
        for item in cronograma:
            item['Valor_Presente'] *= fator_ajuste
            item['Desconto_Aplicado'] = item['Valor'] - item['Valor_Presente']
        
        total_valor = sum(p['Valor'] for p in cronograma)
        total_valor_presente = sum(p['Valor_Presente'] for p in cronograma)
        total_desconto = sum(p['Desconto_Aplicado'] for p in cronograma)
        
        cronograma.append({
            "Item": "TOTAL",
            "Tipo": "",
            "Data_Vencimento": "",
            "Data_Pagamento": "",
            "Dias": "",
            "Valor": total_valor,
            "Valor_Presente": total_valor_presente,
            "Desconto_Aplicado": total_desconto
        })
    
    except Exception as e:
        st.error(f"Erro ao gerar cronograma: {str(e)}")
        cronograma = [{
            "Item": "ERRO",
            "Tipo": "Erro no cálculo",
            "Data_Vencimento": "",
            "Data_Pagamento": "",
            "Dias": "",
            "Valor": 0,
            "Valor_Presente": 0,
            "Desconto_Aplicado": 0
        }]
    
    return cronograma

def gerar_pdf(cronograma, dados):
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        pdf.cell(200, 10, txt="Simulação de Financiamento", ln=1, align='C')
        pdf.ln(10)
        
        pdf.cell(200, 10, txt=f"Valor Total: {formatar_moeda(dados['valor_total'])}", ln=1)
        pdf.cell(200, 10, txt=f"Entrada: {formatar_moeda(dados['entrada'])}", ln=1)
        pdf.cell(200, 10, txt=f"Valor Financiado: {formatar_moeda(dados['valor_financiado'])}", ln=1)
        pdf.cell(200, 10, txt=f"Taxa Mensal: {dados['taxa_mensal']}%", ln=1)
        pdf.ln(10)
        
        pdf.cell(200, 10, txt=f"Comissão Coordenação: {formatar_moeda(dados['comissoes']['coord'])}", ln=1)
        pdf.cell(200, 10, txt=f"Comissão Imobiliária: {formatar_moeda(dados['comissoes']['imob'])}", ln=1)
        pdf.cell(200, 10, txt=f"Total Comissões: {formatar_moeda(dados['comissoes']['total'])}", ln=1)
        pdf.ln(10)
        
        colunas = ["Item", "Tipo", "Data Venc.", "Valor", "Valor Presente"]
        larguras = [30, 30, 40, 40, 40]
        
        for col, larg in zip(colunas, larguras):
            pdf.cell(larg, 10, txt=col, border=1)
        pdf.ln()
        
        for item in [p for p in cronograma if p['Item'] != 'TOTAL']:
            pdf.cell(larguras[0], 10, txt=item['Item'], border=1)
            pdf.cell(larguras[1], 10, txt=item['Tipo'], border=1)
            pdf.cell(larguras[2], 10, txt=item['Data_Vencimento'], border=1)
            pdf.cell(larguras[3], 10, txt=formatar_moeda(item['Valor']), border=1)
            pdf.cell(larguras[4], 10, txt=formatar_moeda(item['Valor_Presente']), border=1)
            pdf.ln()
        
        total = next(p for p in cronograma if p['Item'] == 'TOTAL')
        pdf.cell(sum(larguras[:3]), 10, txt="TOTAL", border=1, align='R')
        pdf.cell(larguras[3], 10, txt=formatar_moeda(total['Valor']), border=1)
        pdf.cell(larguras[4], 10, txt=formatar_moeda(total['Valor_Presente']), border=1)
        
        pdf_output = BytesIO()
        pdf_data = pdf.output(dest='S')
        # Verifica se precisa codificar (para Python 3.x)
        if isinstance(pdf_data, str):
            pdf_data = pdf_data.encode('latin1')
        pdf_output.write(pdf_data)
        pdf_output.seek(0)
        
        return pdf_output
    except Exception as e:
        st.error(f"Erro ao gerar PDF: {str(e)}")
        return BytesIO()

def gerar_excel(cronograma):
    try:
        # Verifica e instala openpyxl se necessário
        try:
            import openpyxl
        except ImportError:
            import subprocess
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl
            
        output = BytesIO()
        
        df = pd.DataFrame([p for p in cronograma if p['Item'] != 'TOTAL'])
        total = next(p for p in cronograma if p['Item'] == 'TOTAL')
        
        df = pd.concat([df, pd.DataFrame([total])], ignore_index=True)
        
        # Formatação de moeda
        for col in ['Valor', 'Valor_Presente', 'Desconto_Aplicado']:
            df[col] = df[col].apply(lambda x: formatar_moeda(x) if pd.notnull(x) else 'R$ 0,00')
        
        # Garante que o ExcelWriter usará openpyxl
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Erro ao gerar Excel: {str(e)}")
        return BytesIO()

def main():
    set_theme()
    
    # Layout com logo e título
    logo = load_logo()
    if logo:
        col1, col2 = st.columns([1, 4])
        with col1:
            st.image(logo, width=150, use_container_width=False)
        with col2:
            st.title("Simulador Imobiliária Celeste")
    else:
        st.title("Simulador Imobiliária Celeste")   
   
    # Inicializa variáveis de sessão
    if 'valor_total' not in st.session_state:
        st.session_state.valor_total = 500000.0
    if 'entrada' not in st.session_state:
        st.session_state.entrada = 50000.0
    if 'valor_parcela' not in st.session_state:
        st.session_state.valor_parcela = 0.0
    if 'valor_balao' not in st.session_state:
        st.session_state.valor_balao = 0.0
    
    with st.form("simulador_form"):
        col1, col2 = st.columns(2)
        
        with col1:
            # Inputs numéricos com formatação manual
            valor_total = st.number_input(
                "Valor Total do Imóvel (R$)", 
                min_value=0.0, 
                value=st.session_state.valor_total, 
                step=1000.0,
                format="%.2f"
            )
            entrada = st.number_input(
                "Entrada (R$)", 
                min_value=0.0, 
                value=st.session_state.entrada, 
                step=1000.0, 
                max_value=valor_total,
                format="%.2f"
            )
            
            data_input = st.date_input("Data de Entrada", value=datetime.now(), format="DD/MM/YYYY")
            data_entrada = datetime.combine(data_input, datetime.min.time())
            taxa_mensal = st.number_input("Taxa de Juros Mensal (%)", min_value=0.0, value=0.79, step=0.01)
            modalidade = st.selectbox(
                "Modalidade de Pagamento",
                options=["mensal", "mensal + balão", "só balão anual", "só balão semestral"],
                index=0
            )
            
            # Seletor de Tipo de Balão (só aparece quando modalidade é "mensal + balão")
            tipo_balao = None
            if modalidade == "mensal + balão":
                tipo_balao = st.selectbox(
                    "Tipo de balão:",
                    options=["Anual", "Semestral"],
                    index=0
                )
            elif modalidade == "só balão anual":
                tipo_balao = "Anual"
            elif modalidade == "só balão semestral":
                tipo_balao = "Semestral"
        
        with col2:
            qtd_parcelas = st.number_input("Quantidade de Parcelas", min_value=1, value=120, step=1)
            
            # Atualiza quantidade de balões baseado na modalidade e tipo de balão
            qtd_baloes = 0
            if modalidade in ["mensal + balão", "só balão anual", "só balão semestral"]:
                qtd_baloes = atualizar_baloes(modalidade, qtd_parcelas, tipo_balao.lower() if tipo_balao else None)
                st.write(f"Quantidade de Balões: {qtd_baloes}")
            
            valor_parcela = st.number_input(
                "Valor da Parcela (R$ - deixe 0 para cálculo automático)", 
                min_value=0.0, 
                value=st.session_state.valor_parcela, 
                step=100.0,
                format="%.2f"
            )
            
            valor_balao = 0.0
            if modalidade in ["mensal + balão", "só balão anual", "só balão semestral"]:
                valor_balao = st.number_input(
                    "Valor do Balão (R$ - deixe 0 para cálculo automático)", 
                    min_value=0.0, 
                    value=st.session_state.valor_balao, 
                    step=1000.0,
                    format="%.2f"
                )
            
            comissao_coordenacao = st.number_input("Comissão de Coordenação (%)", min_value=0.0, value=0.5, step=0.1)
            comissao_imobiliaria = st.number_input("Comissão Imobiliária (%)", min_value=0.0, value=5.0, step=0.1)
        
        # Botão de submit dentro do form
        submitted = st.form_submit_button("Calcular")
    
    if submitted:
        try:
            # Atualiza os valores na sessão
            st.session_state.valor_total = valor_total
            st.session_state.entrada = entrada
            st.session_state.valor_parcela = valor_parcela
            st.session_state.valor_balao = valor_balao
            
            if valor_total <= 0 or entrada < 0:
                st.error("Valor total e entrada são obrigatórios")
                return
            
            valor_financiado = max(valor_total - entrada, 0)
            
            if valor_financiado <= 0:
                st.error("Valor financiado deve ser maior que zero")
                return
            
            taxas = calcular_taxas(taxa_mensal)
            comissoes = calcular_comissoes(valor_total, comissao_coordenacao, comissao_imobiliaria)
            
            modo = determinar_modo_calculo(modalidade)
            
            if modo == 1:  # Somente parcelas mensais
                valor_parcela = calcular_parcela(valor_financiado, taxas['mensal'], qtd_parcelas)
                valor_balao = 0
                qtd_baloes = 0
            elif modo == 2:  # Parcela + balão
                if valor_parcela > 0:
                    saldo_parcelas = calcular_valor_presente_total(valor_parcela, taxas['mensal'], qtd_parcelas)
                    saldo_baloes = max(valor_financiado - saldo_parcelas, 0)
                    valor_balao = calcular_parcela(saldo_baloes, taxas[tipo_balao.lower()], qtd_baloes)
                else:
                    saldo_baloes = calcular_valor_presente_total(valor_balao, taxas[tipo_balao.lower()], qtd_baloes)
                    saldo_parcelas = max(valor_financiado - saldo_baloes, 0)
                    valor_parcela = calcular_parcela(saldo_parcelas, taxas['mensal'], qtd_parcelas)
            elif modo == 3:  # Só balão anual
                valor_balao = calcular_parcela(valor_financiado, taxas['anual'], qtd_baloes)
                valor_parcela = 0
                qtd_parcelas = 0
            elif modo == 4:  # Só balão semestral
                valor_balao = calcular_parcela(valor_financiado, taxas['semestral'], qtd_baloes)
                valor_parcela = 0
                qtd_parcelas = 0
            
            cronograma = gerar_cronograma(
                valor_financiado, 
                valor_parcela, 
                valor_balao,
                qtd_parcelas, 
                qtd_baloes, 
                modalidade, 
                tipo_balao.lower() if tipo_balao else None,
                data_entrada, 
                taxas
            )
            
            # Mostrar resultados
            st.subheader("Resultados da Simulação")
            
            col_res1, col_res2, col_res3 = st.columns(3)
            
            with col_res1:
                st.metric("Valor Total", formatar_moeda(valor_total))
                st.metric("Entrada", formatar_moeda(entrada))
                st.metric("Valor Financiado", formatar_moeda(valor_financiado))
            
            with col_res2:
                st.metric("Taxa Mensal", f"{taxa_mensal}%")
                st.metric("Comissão Coordenação", formatar_moeda(comissoes['coord']))
                st.metric("Comissão Imobiliária", formatar_moeda(comissoes['imob']))
            
            with col_res3:
                st.metric("Total Comissões", formatar_moeda(comissoes['total']))
                st.metric("Valor da Parcela", formatar_moeda(valor_parcela))
                if modalidade != "mensal":
                    st.metric("Valor do Balão", formatar_moeda(valor_balao))
            
            # Mostrar cronograma
            st.subheader("Cronograma de Pagamentos")
            
            df_cronograma = pd.DataFrame([p for p in cronograma if p['Item'] != 'TOTAL'])
            
            # Formatação das colunas numéricas
            df_cronograma['Valor'] = df_cronograma['Valor'].apply(lambda x: formatar_moeda(x))
            df_cronograma['Valor_Presente'] = df_cronograma['Valor_Presente'].apply(lambda x: formatar_moeda(x))
            df_cronograma['Desconto_Aplicado'] = df_cronograma['Desconto_Aplicado'].apply(lambda x: formatar_moeda(x))
            
            st.dataframe(df_cronograma)
            
            total = next(p for p in cronograma if p['Item'] == 'TOTAL')
            st.metric("Valor Total a Pagar", formatar_moeda(total['Valor']))
            st.metric("Valor Presente Total", formatar_moeda(total['Valor_Presente']))
            
            # Botões de exportação
            st.subheader("Exportar Resultados")
            
            col_exp1, col_exp2 = st.columns(2)
            
            with col_exp1:
                pdf_file = gerar_pdf(cronograma, {
                    'valor_total': valor_total,
                    'entrada': entrada,
                    'taxa_mensal': taxa_mensal,
                    'comissoes': comissoes,
                    'valor_financiado': valor_financiado
                })
                st.download_button(
                    label="Exportar para PDF",
                    data=pdf_file,
                    file_name="simulacao_financiamento.pdf",
                    mime="application/pdf"
                )
            
            with col_exp2:
                excel_file = gerar_excel(cronograma)
                st.download_button(
                    label="Exportar para Excel",
                    data=excel_file,
                    file_name="simulacao_financiamento.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        except Exception as e:
            st.error(f"Ocorreu um erro durante a simulação: {str(e)}")

if __name__ == '__main__':
    main()
